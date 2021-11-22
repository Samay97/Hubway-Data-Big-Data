from os import path

from airflow.utils.decorators import apply_defaults
from airflow.models.baseoperator import BaseOperator

from ast import literal_eval

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


# Collums
month_col = 1
avg_trip_dur_col = 2
avg_trip_dis_col = 3
gender_share_col = 4
gender_share_m_col = 4
gender_share_f_col = 5
gender_share_d_col = 6
age_share_col = 7
top_used_bikes_col = 10
top_start_stations_col = 13
top_end_stations_col = 16
time_slots_col = 19

# Rows
month_row_start = 5
avg_trip_dur_row = 5
avg_trip_dis_row = 5
gender_share_row = 5
age_share_row = 5
top_used_bikes_row = 5
top_start_stations_row = 5
top_end_stations_row = 5
time_slots_row = 5

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
gray_color = Color(rgb='00ADADAD')
gray_solid = PatternFill(patternType='solid', fgColor=gray_color)
header_font = Font(size = '16', bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')


class CreateExcelFromCSV(BaseOperator):

    template_fields = ('csv_path', 'excel_path')
    ui_color = '#1F6E43'
    
    @apply_defaults
    def __init__(self, csv_path: str, excel_path: str, **kwargs) -> None:
        super().__init__(**kwargs)
        self.csv_path = csv_path
        self.excel_path = excel_path
    
    def read_csv(self, path) -> pd.DataFrame:
        return pd.read_csv(path)
    
    def create_basic_excel_layout(self, sheets) -> Workbook:
        # create new excel instance
        excel_workbook = Workbook()

        # activate sheet to write in
        sheet = excel_workbook.active
        sheet.cell(row=1, column=1).value = "Hubway - Data from 2015 to 2017: Calculating per month"

        return excel_workbook
    
    def create_sheet_layout(self, workbook: Workbook, year, i) -> Worksheet:
        sheet = workbook.create_sheet('KPI\'S {}'.format(year), index=i)
        sheet.cell(row=1, column=1).value = 'Hubway - KPI from {}: Calculated per month'.format(year)
        
        # Create collums
        sheet.cell(row=3, column=month_col).value = 'Month'
        sheet.cell(row=3, column=avg_trip_dur_col).value = 'AVG trip duration'
        sheet.cell(row=3, column=avg_trip_dis_col).value = 'AVG trip distance'
        sheet.cell(row=3, column=gender_share_col).value = 'Gender Share'
        sheet.cell(row=4, column=gender_share_m_col).value = 'Men'
        sheet.cell(row=4, column=gender_share_f_col).value = 'Female'
        sheet.cell(row=4, column=gender_share_d_col).value = 'Divers'
        sheet.cell(row=3, column=age_share_col).value = 'Top Age Share'
        sheet.cell(row=4, column=age_share_col).value = '1.'
        sheet.cell(row=4, column=age_share_col+1).value = '2.'
        sheet.cell(row=4, column=age_share_col+2).value = '3.'
        sheet.cell(row=3, column=top_used_bikes_col).value = 'Top Used Bikes-IDs'
        sheet.cell(row=4, column=top_used_bikes_col).value = '1.'
        sheet.cell(row=4, column=top_used_bikes_col+1).value = '2.'
        sheet.cell(row=4, column=top_used_bikes_col+2).value = '3.'
        sheet.cell(row=3, column=top_start_stations_col).value = 'Top Start Stations'
        sheet.cell(row=4, column=top_start_stations_col).value = '1.'
        sheet.cell(row=4, column=top_start_stations_col+1).value = '2.'
        sheet.cell(row=4, column=top_start_stations_col+2).value = '3.'
        sheet.cell(row=3, column=top_end_stations_col).value = 'Top End Stations'
        sheet.cell(row=4, column=top_end_stations_col).value = '1.'
        sheet.cell(row=4, column=top_end_stations_col+1).value = '2.'
        sheet.cell(row=4, column=top_end_stations_col+2).value = '3.'
        sheet.cell(row=3, column=time_slots_col).value = 'Trips time-slot usage'
        sheet.cell(row=4, column=time_slots_col).value = '0 - 6 Clock'
        sheet.cell(row=4, column=time_slots_col+1).value = '6 - 12 Clock'
        sheet.cell(row=4, column=time_slots_col+2).value = '12 - 18 Clock'
        sheet.cell(row=4, column=time_slots_col+3).value = '18 - 24 Clock'


        sheet.cell(row=23, column=1).value = 'Age Share per month'
        sheet.cell(row=24, column=1).value = 'Month'
        sheet.cell(row=24, column=2).value = 'Age'
        sheet.cell(row=24, column=3).value = 'Count'

        sheet.cell(row=23, column=5).value = 'Top 10 used Bikes-IDs per month'
        sheet.cell(row=24, column=5).value = 'Month'
        sheet.cell(row=24, column=6).value = 'Bike-ID'
        sheet.cell(row=24, column=7).value = 'Count'

        sheet.cell(row=23, column=12).value = 'Top 10 Start Stations per month'
        sheet.cell(row=24, column=12).value = 'Month'
        sheet.cell(row=24, column=13).value = 'Start Station'
        sheet.cell(row=24, column=14).value = 'Count'

        sheet.cell(row=23, column=16).value = 'Top 10 End Stations per month'
        sheet.cell(row=24, column=16).value = 'Month'
        sheet.cell(row=24, column=17).value = 'End Station'
        sheet.cell(row=24, column=18).value = 'Count'

        return sheet


    def design_sheet(self, sheet: Worksheet):
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=6)
        sheet.merge_cells(start_row=3, start_column=gender_share_col, end_row=3, end_column=gender_share_d_col)
        sheet.merge_cells(start_row=3, start_column=age_share_col, end_row=3, end_column=age_share_col+2)
        sheet.merge_cells(start_row=3, start_column=time_slots_col, end_row=3, end_column=time_slots_col+3)
        sheet.merge_cells(start_row=3, start_column=top_start_stations_col, end_row=3, end_column=top_start_stations_col+2)
        sheet.merge_cells(start_row=3, start_column=top_end_stations_col, end_row=3, end_column=top_end_stations_col+2)
        sheet.merge_cells(start_row=23, start_column=1, end_row=23, end_column=3)
        sheet.merge_cells(start_row=23, start_column=5, end_row=23, end_column=7)
        sheet.merge_cells(start_row=23, start_column=12, end_row=23, end_column=14)
        sheet.merge_cells(start_row=23, start_column=16, end_row=23, end_column=18)
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 5
        sheet.column_dimensions['I'].width = 5
        sheet.column_dimensions['J'].width = 6
        sheet.column_dimensions['K'].width = 6
        sheet.column_dimensions['L'].width = 6
        sheet.column_dimensions['M'].width = 30
        sheet.column_dimensions['N'].width = 30
        sheet.column_dimensions['O'].width = 30
        sheet.column_dimensions['P'].width = 30
        sheet.column_dimensions['Q'].width = 30
        sheet.column_dimensions['R'].width = 30
        sheet.column_dimensions['S'].width = 12
        sheet.column_dimensions['T'].width = 12
        sheet.column_dimensions['U'].width = 12
        sheet.column_dimensions['V'].width = 12

        header_cell = sheet.cell(row=1, column=1)
        header_cell.font = header_font
        header_cell.alignment = header_alignment

        # border and color style
        for y in range(3, 17):
            for x in range(1, time_slots_col+4):
                sheet.cell(row=y, column=x).border = thin_border
        
        for y in range(3, 5):
            for x in range(1, time_slots_col+4):
                sheet.cell(row=y, column=x).fill = gray_solid
        for y in range(23, 25):
            for x in range(1, 4):
                sheet.cell(row=y, column=x).border = thin_border
                sheet.cell(row=y, column=x).fill = gray_solid
        for y in range(23, 25):
            for x in range(5, 8):
                sheet.cell(row=y, column=x).border = thin_border
                sheet.cell(row=y, column=x).fill = gray_solid
        for y in range(23, 25):
            for x in range(12, 15):
                sheet.cell(row=y, column=x).border = thin_border
                sheet.cell(row=y, column=x).fill = gray_solid
        for y in range(23, 25):
            for x in range(16, 19):
                sheet.cell(row=y, column=x).border = thin_border
                sheet.cell(row=y, column=x).fill = gray_solid


    def create_full_data_table(self, sheet, df_array, row_index, month, col1, col2, col3, format=False):
        index = row_index
        if format:
                cell1 = sheet.cell(row=25+index, column=col1)
                cell2 = sheet.cell(row=25+index, column=col1)
                cell3 = sheet.cell(row=25+index, column=col1)
                cell1.number_format = FORMAT_PERCENTAGE_00
                cell2.number_format = FORMAT_PERCENTAGE_00
                cell3.number_format = FORMAT_PERCENTAGE_00
                for i, data in enumerate(df_array):
                    cell1.value = month
                    cell2.value = data[0]
                    cell3.value = data[1]
                    index += 1
        else:
            for i, data in enumerate(df_array):
                    sheet.cell(row=25+index, column=col1).value = month
                    sheet.cell(row=25+index, column=col2).value = data[0]
                    sheet.cell(row=25+index, column=col3).value = data[1]
                    index += 1
        return index
    
    def execute(self, context):
        self.log.info('Create excel from: {}'.format(self.csv_path))

        dataframe = self.read_csv(self.csv_path)
        excel_workbook = Workbook()               
        sorted_df = dataframe.sort_values(['year'], ascending=[True])['year'].unique()


        for i, year in enumerate(sorted_df):
            print('Create KPIS for: ' + str(year))
            sheet = self.create_sheet_layout(excel_workbook, year, i)
            self.design_sheet(sheet)
        
            df = dataframe.query('year == {}'.format(year)).sort_values(['month'], ascending=[True])
            df = df.to_dict()
            
            index = 0
            age_share_index = 0
            top_used_bikes_index = 0
            top_start_stations_index = 0
            top_end_station_index = 0
            for key in df['year'].keys():
                """
                month             
                avg_trip_duration 
                avg_trip_distance 
                gender_share      
                age_share         
                top_used_bikes    
                top_start_stations
                top_end_stations  
                time_slots        
                """
                # Get values from cells
                month = df['month'][key]
                
                avg_trip_duration = df['avg_trip_duration'][key]
                
                avg_trip_distance = df['avg_trip_distance'][key]
                
                # 0-m, 1-w, 2-d
                gender_share = literal_eval(df['gender_share'][key])

                age_share = literal_eval(df['age_share'][key])
                age_share = age_share[:4]

                top_used_bikes = literal_eval(df['top_used_bikes'][key])
                top_used_bikes = top_used_bikes[:3]

                top_start_stations = literal_eval(df['top_start_stations'][key])
                top_start_stations = top_start_stations[:3]

                top_end_stations = literal_eval(df['top_end_stations'][key])
                top_end_stations = top_end_stations[:3]

                time_slots = literal_eval(df['time_slots'][key])

                # set values in sheet
                sheet.cell(row=month_row_start+index, column=month_col).value = month
                sheet.cell(row=avg_trip_dur_row+index, column=avg_trip_dur_col).value = avg_trip_duration
                sheet.cell(row=avg_trip_dis_row+index, column=avg_trip_dis_col).value = avg_trip_distance
                sheet.cell(row=gender_share_row+index, column=gender_share_m_col).value = gender_share[0] /100
                sheet.cell(row=gender_share_row+index, column=gender_share_m_col).number_format = FORMAT_PERCENTAGE_00
                sheet.cell(row=gender_share_row+index, column=gender_share_f_col).value = gender_share[1] /100
                sheet.cell(row=gender_share_row+index, column=gender_share_f_col).number_format = FORMAT_PERCENTAGE_00
                sheet.cell(row=gender_share_row+index, column=gender_share_d_col).value = gender_share[2] /100
                sheet.cell(row=gender_share_row+index, column=gender_share_d_col).number_format = FORMAT_PERCENTAGE_00
                sheet.cell(row=top_used_bikes_row+index, column=top_used_bikes_col).value = top_used_bikes[0][0]
                sheet.cell(row=top_used_bikes_row+index, column=top_used_bikes_col+1).value = top_used_bikes[1][0]
                sheet.cell(row=top_used_bikes_row+index, column=top_used_bikes_col+2).value = top_used_bikes[2][0]
                sheet.cell(row=top_start_stations_row+index, column=top_start_stations_col).value = top_start_stations[0][0]
                sheet.cell(row=top_start_stations_row+index, column=top_start_stations_col+1).value = top_start_stations[1][0]
                sheet.cell(row=top_start_stations_row+index, column=top_start_stations_col+2).value = top_start_stations[2][0]
                sheet.cell(row=top_end_stations_row+index, column=top_end_stations_col).value = top_end_stations[0][0]
                sheet.cell(row=top_end_stations_row+index, column=top_end_stations_col+1).value = top_end_stations[1][0]
                sheet.cell(row=top_end_stations_row+index, column=top_end_stations_col+2).value = top_end_stations[2][0]
                i = 0
                for age in age_share:
                    if age[0] != 0:
                        sheet.cell(row=age_share_row+index, column=age_share_col+i).value = age[0]
                        i += 1
                        if i ==3:
                            break
                for time_slot_index, time_slot in enumerate(time_slots):
                    sheet.cell(row=time_slots_row+index, column=time_slots_col+time_slot_index).number_format = FORMAT_PERCENTAGE_00
                    sheet.cell(row=time_slots_row+index, column=time_slots_col+time_slot_index).value = time_slot[1]/100

                # Create Full Data tabels
                age_share_index = self.create_full_data_table(
                        sheet,
                        literal_eval(df['age_share'][key]), 
                        age_share_index, 
                        month,
                        1,2,3
                )

                top_used_bikes_index = self.create_full_data_table(
                        sheet,
                        literal_eval(df['top_used_bikes'][key]), 
                        top_used_bikes_index, 
                        month,
                        5,6,7
                )
                
                top_start_stations_index = self.create_full_data_table(
                        sheet,
                        literal_eval(df['top_start_stations'][key]),
                        top_start_stations_index,
                        month,
                        12,13,14
                )

                top_end_station_index = self.create_full_data_table(
                        sheet,
                        literal_eval(df['top_end_stations'][key]),
                        top_end_station_index,
                        month,
                        16,17,18
                )        

                index += 1

        # remove default sheet
        excel_workbook.remove(excel_workbook['Sheet'])
        excel_workbook.save(self.excel_path + 'hubway-tripdata.xlsx')

        self.log.info('All Done. Excel file: {}'.format(self.excel_path + 'hubway-tripdata.xlsx'))
